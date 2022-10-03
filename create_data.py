
from sklearn.linear_model import LinearRegression
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import pprint

wb = openpyxl.load_workbook('D:\SUNDAY1\\123.xlsx')
sheet = wb['Sheet1']
# đo⁄n code sinh ra dœ li»u
numOfPoint = 30
noise = np.random.normal(0, 1, numOfPoint).reshape(-1, 1)
x = np.linspace(30, 100, numOfPoint).reshape(-1, 1)
N = x.shape[0]
y = (15)*x + 8 + 20*(noise)
print(y[0])
plt.scatter(x, y)


for i in range(1, numOfPoint+1):
    sheet.cell(row=i, column=1, value=float(x[i-1, 0]))
    sheet.cell(row=i, column=2, value=float(y[i-1, 0]))
    pprint.pprint(list(sheet.values), width=20)
wb.save('D:\SUNDAY1\\123.xlsx')
